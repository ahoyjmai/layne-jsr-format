######################################################################################
# HOW TO USE
# Download latest JSR file from power BI and insert into "input" folder here:
input_dir = "../input new jsr"
input_dir2 = "../input prev jsr"

# Double click the .py script and it will run.
# It will generate a updated service file here:
output_dir = "../output"
#
# PURPOSE
# Reformats JSR downloaded from BI to a format that is more usable for monthly review.
# Calculates values that were not included in original BI download file.
# Eventually will be able to compile a report with information from other sources
#
# OTHER
# JSR file must have "JSR" in name
# This program relies on Power BI output being consistent. any column modifications or
# unexpected additions to input spreadsheets will cause this script to malfunction.
# equations are hard-coded in HEADERMAP
#
# Written by Jonathan Mai on 8/7/2019
Last_Updated = "Last Updated 11/14/2019"
Version_Number = "Version v1.4"

# version Notes

# v1.3
# various directory related bugs fixed. Pyinstaller is able to turn this into EXE file.
# Use cmd to go to the directory, and then enter: pyinstaller --onefile JSRformatter.py
# exe will be placed in dist folder

# Source code available at https://github.com/ahoyjmai/layne-jsr-format.git

# v1.2
# recovered from accidental deletion of conditional formatting, although it does not handle edge cases yet.
# macro section has safeties and predictible jumps. force kill does not yet work. needs to be ctrl-c in main window.
# type 'debug' for debug mode which runs faster
# now redirects from relative directory instead of absolute directory
######################################################################################
print("Loading...")

import datetime
import os
import sys
import keyboard
import signal
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.comments import Comment
from copy import copy
import numpy as np 
from macro import *
from headermap import *

def main():
        DEBUG=False
        #keyboard.add_hotkey('esc', killer)      # Escape will kill program at any time
        #this doesn't actually do anything, all it does is be scary
        with open("readme.txt",'r') as intro:   # print instructions
                print (intro.read())

        print()
        #os.chdir(output_dir)     #change to directory specified in top notes section
        # cwd=os.getcwd()       #prints name of current directory

        ##### This section finds the filenames and directories for use####
        # locate most recent JSR file
        inputJSR=newest_file(input_dir,"JSR")
        print ("Loading:",inputJSR)
        #print ("   "+inputJSR)

        try:
                wb1 = load_workbook(inputJSR)              # primary workbook being read
        except KeyboardInterrupt:
                raise KeyboardInterrupt
        except:
                print ("Error loading JSR. Closing this file in excel will likely fix the problem.")
                input("Press 'ENTER' to close")
                sys.exit()
        finally: 
                ws1 = wb1.active
                ws1.title= "Original_All"
                newJSRtimestamp = get_timestamp_str(inputJSR,ws1)
                ws1['A2'].value = newJSRtimestamp
                


        #load previous jsr
        inputJSR2=newest_file(input_dir2,"JSR")
        print ("Loading:",inputJSR2)
        #print ("   "+inputJSR2)
        try:
                wbprev = load_workbook(inputJSR2)              # primary workbook being read
        except KeyboardInterrupt:
                raise KeyboardInterrupt
        except:
                print ("Error loading JSR. Closing this file in excel will likely fix the problem.")
                input("Press 'ENTER' to close")
                sys.exit()
                
        # find right worksheet in previous
        #print ("previous JSR loaded, now searching for correct sheet")
        if len(wbprev.sheetnames)>1:
                for possible_title in ["Original_All","Sheet1"]:   #acceptable sheet titles for searching, in order of most preferred (at left) to least preferred (at right)
                        if possible_title in wbprev.sheetnames:
                                if wbprev[possible_title]['A3'].value == "Contract Type" and wbprev[possible_title]['B3'].value == "Business Unit Type": #the powerBI JSR format has these values in headers
                                        wsprev = wbprev[possible_title]
                                        break # once you find it stop searching
        elif wbprev[wbprev.sheetnames[0]]['A3'].value == "Contract Type": #the correct format should have 'contract type' in the 3rd line first column
                wsprev = wbprev[wbprev.sheetnames[0]]
        else:
                print ("Unexpected workbook format, not sure which sheet to use for PREVIOUS MONTH")
                input("Press 'ENTER' to close")
                sys.exit()

        oldJSRtimestamp = get_timestamp_str(inputJSR2,wsprev)
        wsprev['A2'].value = oldJSRtimestamp #not this one because it would be putting the value in the file in "old JSR" isntead of the newly generated one.
        print()
        print("      New JSR dated", newJSRtimestamp)
        print("     Prev JSR dated", oldJSRtimestamp)
        print()
        print("Please note: To kill this script at any time, press Ctrl-C in this window")
        print()        
        print("If these files & dates look good, PRESS 'ENTER' TO START")
                
        start=input()
        
        if start.lower()=="debug":
                DEBUG=True
                print("ACTIVATING DEBUG MODE: alternate row highlighting is disabled to let script run faster.")

        print ("--------------------------------------------------------------------------")
        print()
        print ("Initializing all the new worksheets")

        #create all the new worksheets
        ws2 = newSheetWithHeaders(wb1,"All Areas",HEADERMAP)
        regional_worksheet_list = [
                newWorksheet("West591","591",wb1),
                newWorksheet("Southwest592","592",wb1),
                newWorksheet("Central586","586",wb1),
                newWorksheet("SouthEast587","587",wb1),
                newWorksheet("NorthEast588","588",wb1),
                newWorksheet("Treatment590","590",wb1),
                ]
        wsother = newSheetWithHeaders(wb1,"Other",HEADERMAP)

        # setting up a list of all worksheets for convenience
        all_modified_worksheet_list = []
        all_modified_worksheet_list.append(ws2)
        for sheet in regional_worksheet_list:
                all_modified_worksheet_list.append(sheet.body)
        all_modified_worksheet_list.append(wsother)        



        # mapping data in ws1 to new format in ws2
        print ("Transferring original values into re-mapped and re-calculated spreadsheet format")

        # load data from New JSR into python array
        firstRow = 4
        firstCol = 1
        nCols = ws1.max_column
        nRows = ws1.max_row
        if nCols!=68 : print("Possible error, expected New JSR to have 68 columns but instead there is:",nCols)
        allCells = np.array([[cell.value for cell in row] for row in ws1.iter_rows()])
        data = allCells[(firstRow-1):(firstRow-1+nRows),(firstCol-1):(firstCol-1+nCols)]
        data = data[data[:,3].argsort()]       # sort secondarily by job number
        data = data[::-1] #reverse for increasing job numbers (isntead of decreasing)
        data = data[data[:,1].argsort(kind="stable")]       # sort primarily by contract type
        data = data[::-1] #reverse for Closed jobs at end (instead of beginning)


        # load data from Prev JSR into python array
        firstRow = 4
        firstCol = 1
        nCols = wsprev.max_column
        nRows = wsprev.max_row
        if nCols!=68 : print("Possible error, expected Prev JSR to have 68 columns but instead there is:",nCols)
        allCells = np.array([[cell.value for cell in row] for row in wsprev.iter_rows()])
        prevdata = allCells[(firstRow-1):(firstRow-1+nRows),(firstCol-1):(firstCol-1+nCols)]

        for i,row in enumerate(data,1):
                if i%500==0:
                        print("... on row",i,"of",ws1.max_row)
                #if i > 210:      # For debugging, only do 210 rows since this takes a while
                 #       break    # comment this section out when live

                trythisrowfirst=0       #speeds up vlookup by searching this row first, which should be the search hit from the last time.
                for j,col in enumerate(HEADERMAP,1):
                        calc=0
                        if col[2]!="" or col[3]!="":    #interpreting headermap add/subtract terms
                                if col[2]!="":
                                        addvalue=row[xcol(col[2])]
                                        if addvalue:          # if this exists, add the base value
                                                calc = addvalue
                                if col[3]!="":
                                        if col[4]=="prev":                            # if 4th var is "prev" use wsprev instead of ws1. need to lookup by job number
                                                #i=lookup job number
                                                #print("now on newsheet row ",i," trying to complete col ",j,", but we need to do a vlookup")
                                                #print("starting vlookup...",end="")
                                                for k in [trythisrowfirst]+list(range(1,len(prevdata))):
                                                        lookupjobnum=row[3]
                                                        if prevdata[k][3] == lookupjobnum:     # if job numbers match
                                                                subtractvalue=prevdata[k][xcol(col[3])]
                                                                if subtractvalue is not "" and subtractvalue is not None:        #just in case of var type error
                                                                        calc = calc - subtractvalue
                                                                        #print ("subtracting", subtractvalue)
                                                                #else:
                                                                 #       print ("job",lookupjobnum,"col",j,": previous data did not have expected datatype")
                                                                        #input()
                                                                trythisrowfirst=k
                                                                break
                                        else:                                  # determine if we need to get from previous or current JSR. if no 3rd var, just subtract from ws1
                                                subtractvalue=row[xcol(col[3])]               # subtract the second value
                                                if subtractvalue is not "" and subtractvalue is not None:        #just in case of var type error
                                                        calc = calc - subtractvalue
                                                        
                                ws2.cell(row=i+3,column=j).value = calc
                                #ws2.cell(row=i,column=j).number_format = copy(ws1[col[1] + str(i)].number_format)       # copy over the formatting

        ##########################################################################
        ######### CONDITIONAL FORMATTING AND SPECIALTY CALCULATIONS GO HERE#######
        ##########################################################################                        
        #
        #This is where we do special rules for worksheet modified_all
        #

        fudge = 10  # fudge number in case of rounding errors
        print ("Adding conditional formatting and special if-then calculations")
        for i,row in enumerate(ws2.iter_rows(),1):
                if i <= 3:      # skip the first 3 rows
                        continue
                mark_if_actual_cost_is_greater_than_forecasted_cost(row)
                clean_sales_vs_billings_values(row)
                mark_large_POC_receivables(row)
                mark_billings_over_contract_value(row)
                mark_actual_cost_over_billings_by_a_lot(row)
                add_number_formatting(row,HEADERMAP)
                #mark_billings_not_proportional_to_Cost(row)
                
        #
        #        
        #
        #
        ##########################################################################
        ## END OF CONDITIONAL FORMATTING AND SPECIALTY CALCULATIONS###############
        ##########################################################################

        print ("Now splitting modified spreadsheet into regional sheets")

        #split data from newly mapped sheet into multiple regional sheets
        for i,row in enumerate(ws2.iter_rows(),1):
                if i <= 3:      # skip the first 3 rows
                        continue
                if row[xcol("AU")].value is None:                   # Col AU contains the region code
                        continue
                if i%500==0:
                                print("... on row",i,"of",ws2.max_row)

                putinother=True        
                for worksheet in regional_worksheet_list:
                        #print (" ", end="")
                        if worksheet.code in row[xcol("AU")].value:             # copy to regional sheet if code matches
                                nextrow=worksheet.body.max_row+1
                                for cell in row:
                                        new_cell = worksheet.body.cell(row=nextrow, column=cell.col_idx, value= cell.value)
                                        if cell.has_style:
                                                new_cell.border = copy(cell.border)
                                                new_cell.fill = copy(cell.fill)
                                                new_cell.number_format = copy(cell.number_format)
                                                new_cell.comment = copy(cell.comment)
                                                #new_cell.protection = copy(cell.protection)
                                                #new_cell.alignment = copy(cell.alignment)
                                putinother=False                                        
                                break
                if putinother is True:                # copy to Other page if no match
                        nextrow=wsother.max_row+1
                        #print (row[xcol("AU")].value,"- Other")
                        for cell in row:
                                new_cell = wsother.cell(row=nextrow, column=cell.col_idx, value= cell.value)
                                if cell.has_style:
                                        new_cell.border = copy(cell.border)
                                        new_cell.fill = copy(cell.fill)
                                        new_cell.number_format = copy(cell.number_format)
                                        new_cell.comment = copy(cell.comment)

        print ("Now preparing highlighted rows & styling")

        #this takes a while to execute so hide comment it out when running if you're not specifically testing it 
        for sheet in all_modified_worksheet_list:
                #print("Highlighting alternate rows")
                if not DEBUG:
                        highlight_alternate_rows(sheet)        # option to skip this during debug mode
                
                #print("Hiding TBD columns")
                for col in ['B', 'Q', 'R', 'S', 'AT', 'AV', 'BB']:
                        sheet.column_dimensions[col].hidden= True
                #print("Hiding first 2 blank rows")
                # actually, only hiding row 2. Row 1 is left for spreadsheet title.
                for row in [1]:
                        sheet.row_dimensions[row].hidden= True
                # freezing pane
                sheet.freeze_panes = sheet['A4']
                #print("Adjusting Column Widths")
                for i in range(46,57):
                        sheet.column_dimensions[get_column_letter(i)].width = 15
                for i in list(range(6,34)) + list(range(36,46)) + list(range(50,53)):
                        sheet.column_dimensions[get_column_letter(i)].width = 12
                sheet.column_dimensions['E'].width = 35
                
                #print("Putting Filter in Place")
                FullRange = "A3:" + get_column_letter(sheet.max_column)  + str(sheet.max_row)
                sheet.auto_filter.ref = FullRange

        # move original sheet to end 
        move_sheet(wb1,0,len(wb1._sheets)-1)

        # format and include prev month sheet
        prevworksheet = wb1.create_sheet(title="Original_Prev_All")
        for row in wsprev.iter_rows():
                prevworksheet.append(cell.value for cell in row)
        for cell in prevworksheet[3]:       # formats with colors and wraptext
                cell.fill = PatternFill(start_color='5B9BD5',fill_type = "solid")   # start_color is background color, end_color is font color
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(color="FFFFFF",bold=True)
        if not DEBUG:
                highlight_alternate_rows(prevworksheet)
        
        prevworksheet['B2'].value = oldJSRtimestamp

        ws2.active = 0
        ws1.views.sheetView[0].tabSelected = False

        # create save file name based on todays year-month-date
        # this is the filename you save the formatted JSR as
        YMD = datetime.datetime.now().strftime("%y%m%d")
        HMS = datetime.datetime.now().strftime("%H%M%S")

        save_directory = output_dir+"/Reports "+datetime.datetime.now().strftime("%Y-%m-%d")+"/" #ends in /"
        save_file_name = "JSR "+newJSRtimestamp+" vs "+oldJSRtimestamp+"("+YMD+"-"+HMS+")"+".xlsx"
        savefile=save_directory + save_file_name

        print ("")
        print ("Excel processing finished.")
        
        try:
                if not os.path.isdir(save_directory):
                        os.makedirs(save_directory)
                print ("Attempting to save to: ")
                print (savefile)
                wb1.save(savefile)
                print("Save complete")
        except KeyboardInterrupt:
                raise KeyboardInterrupt
        except:
                print("----------------------------------------------------------")
                print("ERROR: Could not save. Close the destination file and then")
                input("Press 'ENTER' to try saving again")
                try:
                        print("Trying last time to save... ",end="")
                        wb1.save(savefile)
                        print("Save complete")
                except KeyboardInterrupt:
                        raise KeyboardInterrupt
                except:
                        print("ERROR: Still could not save. Terminating script.")
                        return wb1
                        sys.exit(0)
                        
        print()
        #print("savefile =",savefile)
        #print("save_file_name",save_file_name)
        AUTOMATE_EXCEL_FORMATTING(savefile,save_file_name)
        input("Script complete. Press 'ENTER' to close")

def newSheetWithHeaders(workbook,sheettitle,headermap):
        worksheet = workbook.create_sheet(title=sheettitle)     # create new sheet with region title
        
        newheaders=[]
        oldheaders=[]
        oldws = workbook["Original_All"]
        for col in headermap:           # extracts and creates list of header names from headermap
                newheaders.append(col[0])
                if col[2]!="" and col[3]=="":
                        oldheaders.append(oldws[col[2]+'3'].value)
                else:   oldheaders.append("")
                                
        worksheet.append(list())        #add 2 blank rows
        worksheet.append(oldheaders)
        worksheet.append(newheaders)       # adds header names to worksheet
        
        for cell in worksheet[3]:       # formats with colors and wraptext
                cell.fill = PatternFill(start_color='5B9BD5',fill_type = "solid")   # start_color is background color, end_color is font color
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(color="FFFFFF",bold=True)
                
        return worksheet

                
class newWorksheet:
        def __init__(self, title, code, workbook):
                self.title=title
                self.code=code
                self.body= newSheetWithHeaders(workbook,title,HEADERMAP)

def move_sheet(wb, from_loc=None, to_loc=None):
        sheets=wb._sheets

        # if no from_loc given, assume last sheet
        if from_loc is None:
                from_loc = len(sheets) - 1

        #if no to_loc given, assume first
        if to_loc is None:
                to_loc = 0

        sheet = sheets.pop(from_loc)
        sheets.insert(to_loc, sheet)


def highlight_alternate_rows(worksheet):
        borderstyle =  Border(top=Side(style='thin',color="9BC2E6"), bottom=Side(style='thin',color="9BC2E6"))
        fillstyle =    PatternFill(start_color='ddebf7',fill_type = "solid")
        
        for i in range (5,worksheet.max_row,2):    # skip first 3(or 4) lines
                worksheet.row_dimensions[i].fill = fillstyle
                worksheet.row_dimensions[i].border = borderstyle
        
        for i in range (4,worksheet.max_row):    # skip first 3(or 4) lines
                if i%100==0:
                        #print(".",end="")
                        sys.stdout.write(".")
                        sys.stdout.flush()
                for cell in worksheet[i]:
                        if i%2==1:  # on odd rows and cells without default fill
                                if cell.fill.start_color.rgb is "00000000":
                                    cell.fill = fillstyle
                                    cell.border = borderstyle
                        if cell.value==0:
                                cell.font = Font(color="b2b2b2")
                        pass                        

def killer():
        print ("killer!")
        signal.CTRL_C_EVENT
        raise KeyboardInterrupt

def xcol(alphanumeric):
        return column_index_from_string(alphanumeric)-1
        
def newest_file(path,keyword=""):
# This function returns the most recently updated file in directory "path"
# Optional, add a filter keyword filenames
        files = os.listdir(path)
        if files ==[]:
                print ("Error: There are no files in",path)
                print ("Please put JSR files into the input folders and try again.")
                input("Press 'ENTER' to quit")
                sys.exit()                       

        tracksheets=[]
        for basename in files:
                if basename.find(keyword)!=-1 :
                        #print (basename)
                        #print (basename.replace('\\','/'))
                        tracksheets.append(basename)
        paths = [os.path.join(path, basename) for basename in tracksheets]
        return max(paths, key=os.path.getctime).replace('\\','/')

def parse_for_date(searchterm,key):
        return searchterm[searchterm.find(key)+len(key):]

def get_timestamp_str(filepath, worksheet):
        # Compares 3 timestamps and returns the earliest one
        #    1. date in cell BP4
        #    2. date file last modified
        #    3. today's date
        # Returns string, ex: '2019-12-31'
      
        strdatefromfilemod = time.strftime('%Y-%m-%d', time.gmtime(os.path.getmtime(filepath)))
        strdatefromexcel = parse_for_date(searchterm=worksheet['BP4'].value,key="||")
        strdatestampnow = datetime.datetime.now().strftime("%Y-%m-%d")  
        
        a= min(strdatefromfilemod, strdatefromexcel, strdatestampnow)
        #print(a, "appears to be the timestamp for", filepath)
        return a

Redfillstyle = PatternFill(start_color='FF7D7D',fill_type = "solid")
Pinkfillstyle = PatternFill(start_color='FFAFAF',fill_type = "solid")

def mark_if_actual_cost_is_greater_than_forecasted_cost(row):
        # O is actual cost, N is Forecasted cost
        if row[xcol("O")].value and row[xcol("M")].value:
                ActualGreaterThanForecast = row[xcol("O")].value - row[xcol("M")].value
                if ActualGreaterThanForecast > 0:
                        row[xcol("O")].fill = Redfillstyle
                        commenttext= "Actual Cost greater than Forecast Cost"
                        row[xcol("O")].comment = Comment(commenttext,"JMai")
                return True
        else: return False
        
def clean_sales_vs_billings_values(row):
        # AC and AD is billings vs sales, delete if negative
        for a in ["AC","AD"]:
                b=row[xcol(a)].value
                if b<0: row[xcol(a)].value =""
        return        
                
def mark_large_POC_receivables(row):
        #Z is POC receivables
        POC_threshold=20000  # ignore unless POC is over $20k
        a= row[xcol("Z")].value
        if a:
                if a>POC_threshold:
                        commenttext= "POC Receivables above $20k"
                        row[xcol("Z")].comment = Comment(commenttext,"JMai")
                        row[xcol("Z")].fill = Redfillstyle
                        return True
                elif a<-POC_threshold:
                        commenttext= "POC Receivables below -$20k"
                        row[xcol("Z")].comment = Comment(commenttext,"JMai")
                        row[xcol("Z")].fill = Redfillstyle
                        return True
        return False
        
def mark_billings_over_contract_value(row):

        # Two levels of highlight. Light Red level 1, Dark Red level 2
        # Level 1: mark if est sales < actual cost * 1.22
        # Level 2: mark if est sales < billings (This one is more important, highlight this one if you have to choose)
        threshold=100
        
        # Y is billings, F is contract Value, O is actual Total Cost
        if row[xcol("Y")].value and row[xcol("F")].value:
                if row[xcol("F")].value > 5:
                        #ignore if contract value is tiny
                        if row[xcol("Y")].value > row[xcol("F")].value + threshold:
                                commenttext = "Billings exceed Contract Value. Change order needed."
                                row[xcol("F")].comment = Comment(commenttext,"JMai")
                                row[xcol("F")].fill = Redfillstyle
                                return True
                        elif row[xcol("O")].value *1.22 > row[xcol("F")].value and row[xcol("O")].value > row[xcol("F")].value + threshold:
                                #apply level 2 only if it is a CJ job
                                if row[xcol("A")].value == "CJ":
                                        commenttext = "Revenue accrual based on actual costs is lower than an 18% margin. Possible CO needed."
                                        row[xcol("F")].comment = Comment(commenttext,"JMai")
                                        row[xcol("F")].fill = Pinkfillstyle
                                        return True
        else: return False

def mark_actual_cost_over_billings_by_a_lot(row):
        
        # Two levels of highlight. Light Red level 1, Dark Red level 2
        # Level 1: mark if cost * 125% > billings
        # Level 2: mark if cost > billings + 15000 (This one is more important, highlight this one if you have to choose)
        
        # Y is billings, O is actual Total Cost
        Cost_Threshold=15000 #ignore unless cost is over billings by a lot, otherwise everything gets flagged
        Cost_Perc_Threshold=1.25
        
        if row[xcol("Y")].value and row[xcol("O")].value:
                if row[xcol("O")].value > 3000 :    # ignore if cost is tiny, less than 3000
                        if row[xcol("O")].value > row[xcol("Y")].value + Cost_Threshold:
                                commenttext = "Actual Cost exceeds Total Billings by over $15k"
                                row[xcol("Y")].comment = Comment(commenttext,"JMai")
                                row[xcol("Y")].fill = Redfillstyle
                                return True
                        elif row[xcol("O")].value * Cost_Perc_Threshold > row[xcol("Y")].value :
                                commenttext = "Billings to Cost Ratio below 1.25"
                                row[xcol("Y")].comment = Comment(commenttext,"JMai")
                                row[xcol("Y")].fill = Pinkfillstyle
                                return True
        return False

def add_number_formatting(row,headermap):
        for i,cell in enumerate(row):
                cell.number_format = headermap[i][1]
        
def mark_billings_not_proportional_to_Cost(row):
        pass


###############################################
#This executes main script
if __name__ == '__main__':
        #try: main()
        main()
#        except (KeyboardInterrupt):
                #print ('Script interrupted by keyboard press')
                #input ("Press enter to close script")
                #sys.exit()



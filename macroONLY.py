print("Loading...")

import datetime
import os
import sys
import keyboard
import signal
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.comments import Comment
from copy import copy
import numpy as np 
from macro import *
from headermap import *


def main():

        print ("Press 'ENTER' to run the macro. Press Ctrl-C to cancel.")
        proceed=input()
        
        if proceed=="":
                print ("YOU HAVE 10 SECONDS TO MAKE SURE JSR EXCEL WINDOW IS IN FRONT AND RECEIVING KEYBOARD COMMANDS.")
                print ("Macro starts in: ",end="")
                time.sleep(1)
                for cnt in range(10):
                        sys.stdout.write(str(10-cnt)+',')
                        sys.stdout.flush()
                        time.sleep(1)

                print ()
                print ("--------------------------------------------------------------------------")
                print ("MACRO STARTED.")
                print ("PLEASE DO NOT USE MOUSE AND KEYBOARD.")
                KEYBOARD_MACRO_START()
                
        else:
                print("You have chosen to cancel the macro.")
        print()
        print("You can use the keyboard and mouse now.")


main()

print("Loading...")
import datetime
import os
import sys
import signal
import time
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment
from copy import copy
import numpy as np
from headermap import HEADERMAP
from macro import *

DEBUG = False
CALC995 = ""

input_dir = "../input new jsr"
input_dir2 = "../input prev jsr"
input_dir3 = "../input MH"
output_dir = "../output"

""""
HOW TO USE
Download latest JSR file from power BI and insert into "input" folders.
Obtain weekly MH report from adam and insert here too.
Double click the .py script and it will run. and create a file into the output folder.

PURPOSE
Reformats JSR downloaded from BI to a format that is more usable for monthly review.
Calculates values that were not included in original BI download file.
Eventually will be able to compile a report with information from other sources

OTHER
JSR file must have "JSR" in name
This program relies on Power BI output being consistent. any column modifications or
unexpected additions to input spreadsheets will cause this script to malfunction.
equations are hard-coded in HEADERMAP

version Notes
v1.8d
 Added 995 total column and 995 total margin.  Fix issue where only one 995 area region supervision was expected in
 MH report, now it sums multiple area region codes if present. 995 MH are negative now. Fixed bugs regarding 995 calc
 w MH. In alternate 995 calculation (for month end) only manhours and est accruals are calculated. N/A is assigned to
 estimated month, estimated total, and estimated margin.
 added flag for forecast margin monthly change

v1.7
 fixed 995 calcs, user has option to skip 995 calcs or col R. Changed regions to Central/NE/SE/Central, eliminated SW

v1.6
 First version allowing 995 calcs. requires two new input files to put into a folder from adam. a YTD MH report
 showing the week end date in filename, and a key file containing each regions 995 rates. calculated 995 cost
 for job numbers only. now it calculates 995 cost for 995 job numbers ie 6009995, with negative value in accruals

v1.5
 Fixed major bug. Changed highlighting rules to use ".value is not None" rather than just ".value"
 This was skipping rules when the value was = zero. especially bad if ignoring cost overruns when billing=0

v1.3
 various directory related bugs fixed. Pyinstaller is able to turn this into EXE file.
 Use cmd to go to the directory, and then enter: pyinstaller --onefile JSRformatter.py
 exe will be placed in dist folder

v1.2
 recovered from accidental deletion of conditional formatting, although it does not handle edge cases yet.
 macro section has safeties and predictible jumps. force kill does not yet work. needs to be ctrl-c in main window.
 type 'debug' for debug mode which runs faster
 now redirects from relative directory instead of absolute directory

Source code available at https://github.com/ahoyjmai/layne-jsr-format.git

Written by Jonathan Mai on 8/7/2019
"""
Last_Updated = "Last Updated 7/30/2020"
Version_Number = "Version v1.8c"


def main():
    # keyboard.add_hotkey('esc', killer)      # Escape will kill program at any time
    # this doesn't actually do anything, all it does is be scary
    with open("readme.txt", 'r') as intro:  # print instructions
        print(intro.read())
    print(Last_Updated)
    print(Version_Number)
    print()

    ##### This section finds the filenames and directories for use####

    # locate and load most recent JSR file
    inputJSR = newest_file(input_dir, "JSR")
    wb1 = trytoloadworkbook(inputJSR)

    ws1 = wb1.active
    ws1.title = "Original_All"
    newJSRtimestamp = get_timestamp_str(inputJSR, ws1)
    ws1['A2'].value = newJSRtimestamp

    # locate and load previous jsr
    inputJSR2 = newest_file(input_dir2, "JSR")
    wbprev = trytoloadworkbook(inputJSR2)

    # find right worksheet in previous
    if len(wbprev.sheetnames) > 1:
        for possible_title in ["Original_All", "Sheet1"]:
            # acceptable sheet titles for searching, in order of most preferred (at left) to least preferred (at right)
            if possible_title in wbprev.sheetnames:
                if wbprev[possible_title]['A3'].value == "Contract Type" and wbprev[possible_title][
                    'B3'].value == "Business Unit Type":
                    # the powerBI JSR format has these values in headers
                    wsprev = wbprev[possible_title]
                    break  # once you find it stop searching
    elif wbprev[wbprev.sheetnames[0]]['A3'].value == "Contract Type":
        # the correct format should have 'contract type' in the 3rd line first column
        wsprev = wbprev[wbprev.sheetnames[0]]
    else:
        print("Unexpected workbook format, not sure which sheet to use for PREVIOUS MONTH")
        input("Press 'ENTER' to close")
        sys.exit()

    # locate and load manhours report
    global CALC995
    inputMHreport = newest_file(input_dir3, "manhours report by allocation area")
    if not inputMHreport:
        CALC995 = "SKIPALL995"
    else:
        wbmh = trytoloadworkbook(inputMHreport)
        # find right worksheet in mh report
        if len(wbmh.sheetnames) > 1:
            for possible_title in ["manhours report by allocation a", "manhours report by allocation", "Sheet1"]:
                # acceptable sheet titles for searching, in order of most preferred (at left) to least preferred (at right)
                if possible_title in wbmh.sheetnames:
                    if wbmh[possible_title]['A1'].value == "ALLOCATION AREA" and wbmh[possible_title][
                        'H1'].value == "SUBSIDIARY":  # If these headers are here, this is the right sheet
                        wsmh = wbmh[possible_title]
                        break  # once you find it stop searching
        elif wbmh[wbmh.sheetnames[0]]['A1'].value == "ALLOCATION AREA":
            # the correct format should have 'contract type' in the 3rd line first column
            wsmh = wbmh[wbmh.sheetnames[0]]
        else:
            print("Could not find correct sheet in Manhours Report")
            CALC995 = "SKIPALL995"

    # locate and load 995 key
    input995key = newest_file(input_dir3, "995 key")
    if not input995key:
        CALC995 = "SKIPALL995"
    else:
        wbkey = trytoloadworkbook(input995key)
        # find right worksheet in key
        if len(wbkey.sheetnames) > 1:
            for possible_title in ["995 Key", "995 key", "key", "KEY", "Key", "Sheet1"]:
                # acceptable sheet titles for 995keys, most preferred (left) to least preferred (right)
                if possible_title in wbkey.sheetnames:
                    if wbkey[possible_title]['C1'].value == "Cost Cntr Home" and wbkey[possible_title]['G1'].value == "Total Rate":
                        # If these headers are here, this is the right sheet
                        wskey = wbkey[possible_title]
                        break  # once you find it stop searching
        elif wbkey[wbkey.sheetnames[0]][
            'G1'].value == "Total Rate":  # the correct format should have 'contract type' in the 3rd line first column
            wskey = wbkey[wbkey.sheetnames[0]]
        else:
            print("Could not find correct sheet in 995 Key")
            CALC995 = "SKIPALL995"

    # Get timestamps from files to verify with user
    oldJSRtimestamp = get_timestamp_str(inputJSR2, wsprev)
    wsprev['A2'].value = oldJSRtimestamp
    if CALC995 != "SKIPALL995":
        mhreporttimestamp = inputMHreport[-12:-5]  # try to get "WE 0419" out of "..... WE 0419.xlsx"
        keytimestamp = time.strftime('%Y-%m-%d', time.gmtime(os.path.getmtime(input995key)))
    else:
        mhreporttimestamp = ""

    print()
    print("      New JSR dated", newJSRtimestamp)
    print("     Prev JSR dated", oldJSRtimestamp)
    if CALC995 != "SKIPALL995":
        print("    MH Report dated", mhreporttimestamp)
        print("      995 Key dated", keytimestamp)
    print()
    print("If these files & dates look good, PRESS 'ENTER' TO START")
    start = input()

    global DEBUG
    if start.lower() == "debug":
        DEBUG = True
        print("ACTIVATING DEBUG (FAST) MODE: 70% of rows are skipped.\n",
              "Alternate row highlighting is disabled to let script run faster.")

    if DEBUG:
        if CALC995 != "SKIPALL995":
            CALC995 = "NORMAL"
    elif CALC995 != "SKIPALL995":
        print("How do you want to handle 995 calculations?\n"
              "  Type '1' for    : Do not perform any 995 calculations (Blank Col R, S, T, U, V)\n"
              "  Type '2' for    : Only do manhours (U) & estimated accrual (V). For month end when 995 costs are already posted.\n"
              "  Type nothing    : Perform all 995 calculations normally\n"
              "Make a selection and press ENTER.")
        CALC995 = input()
        if '1' in CALC995:
            CALC995 = "SKIPALL995"
        elif '2' in CALC995:
            CALC995 = "SKIP995MONTH"
        else:
            CALC995 = "NORMAL"
    else:
        print("Manhours files weren't found so 995 calculations will be skipped (Blank Columns R,S,T,U,V)\n"
             "Press ENTER to continue.")
        input()

    print("--------------------------------------------------------------------------")
    print()
    print("To kill this script at any time, press Ctrl-C in this window")
    print()
    print("Initializing all the new worksheets")

    # create all the new worksheets
    ws2 = newsheetwithheaders(wb1, "All Areas", HEADERMAP, mhreporttimestamp)

    """
    regional_worksheet_list = [
        newWorksheet("West591", "591", wb1, mhreporttimestamp),
        newWorksheet("Southwest592", "592", wb1, mhreporttimestamp),
        newWorksheet("Central586", "586", wb1, mhreporttimestamp),
        newWorksheet("SouthEast587", "587", wb1, mhreporttimestamp),
        newWorksheet("NorthEast588", "588", wb1, mhreporttimestamp),
        newWorksheet("Treatment590", "590", wb1, mhreporttimestamp),
    ]
    """
    central_costctrs = ['KANSAS CITY', 'OMAHA', 'WICHITA', 'GUTHRIE', 'DENVER']
    southeast_costctrs = ['FT. MYERS', 'STUTTGART', 'MEMPHIS', 'RAYNE', 'PENSACOLA',
                          'PRAIRIEVILLE', 'ALBANY', 'SAVANNAH', 'JACKSON', 'HOUSTON',
                          'MIDLAND', 'PLEASANTON']
    northeast_costctrs = ['AURORA', 'ST. LOUIS', 'LONG ISLAND', 'COLLECTOR WELLS', 'BEVERLY',
                          'SCHOHARIE', 'MIDDLETOWN', 'LOUISVILLE', 'WAUSAU', 'HEAVY CIVIL']
    west_costctrs = ['CHANDLER', 'HANFORD', 'REDLANDS', 'WATER TREATMENT']

    regional_worksheet_list = [
        newWorksheet("West", west_costctrs, wb1, mhreporttimestamp, 'red'),
        newWorksheet("Central", central_costctrs, wb1, mhreporttimestamp, 'yellow'),
        newWorksheet("SouthEast", southeast_costctrs, wb1, mhreporttimestamp, 'green'),
        newWorksheet("NorthEast", northeast_costctrs, wb1, mhreporttimestamp, 'blue'),
    ]

    wsother = newsheetwithheaders(wb1, "Other", HEADERMAP, mhreporttimestamp)

    # setting up a list of all worksheets for convenience
    all_modified_worksheet_list = [ws2]
    for sheet in regional_worksheet_list:
        all_modified_worksheet_list.append(sheet.body)
    all_modified_worksheet_list.append(wsother)

    # mapping data in ws1 to new format in ws2
    print("Transferring original values into re-mapped and re-calculated spreadsheet format")

    # load data from New JSR into python array
    firstRow = 4
    firstCol = 1
    nCols = ws1.max_column
    nRows = ws1.max_row
    if nCols != 68:
        print("Possible error, expected New JSR to have 68 columns but instead there is:", nCols)
    allCells = np.array([[cell.value for cell in row] for row in ws1.iter_rows()])
    data = allCells[(firstRow - 1):(firstRow - 1 + nRows), (firstCol - 1):(firstCol - 1 + nCols)]
    data = data[data[:, 3].argsort()]  # sort secondarily by job number
    data = data[::-1]  # reverse for increasing job numbers (isntead of decreasing)
    data = data[data[:, 1].argsort(kind="stable")]  # sort primarily by contract type
    data = data[::-1]  # reverse for Closed jobs at end (instead of beginning)

    # load data from Prev JSR into python array
    firstRow = 4
    firstCol = 1
    nCols = wsprev.max_column
    nRows = wsprev.max_row
    if nCols != 68:
        print("Possible error, expected Prev JSR to have 68 columns but instead there is:", nCols)
    allCells = np.array([[cell.value for cell in row] for row in wsprev.iter_rows()])
    prevdata = allCells[(firstRow - 1):(firstRow - 1 + nRows), (firstCol - 1):(firstCol - 1 + nCols)]

    for i, row in enumerate(data, 1):
        if i % 500 == 0:
            print("... on row", i, "of", ws1.max_row)

        if DEBUG:
            if i % 10 > 3:
                continue  # continue will skip rows numbered 3-9, skipping 70% of the entries.

        trythisrowfirst = 0  # speeds up vlookup by searching this row first, which should be the search hit from the last time.
        for j, col in enumerate(HEADERMAP, 1):
            calc = 0
            if col[2] != "" or col[3] != "":  # interpreting headermap add/subtract terms
                if col[2] != "":
                    addvalue = row[xcol(col[2])]
                    if addvalue:  # if this exists, add the base value
                        calc = addvalue
                if col[3] != "":
                    if col[4] == "prev":  # if 4th var is "prev" use wsprev instead of ws1. need to lookup by job number
                        # i=lookup job number
                        # print("now on newsheet row ",i," trying to complete col ",j,", but we need to do a vlookup")
                        # print("starting vlookup...",end="")
                        for k in [trythisrowfirst] + list(range(1, len(prevdata))):
                            lookupjobnum = row[3]
                            if prevdata[k][3] == lookupjobnum:  # if job numbers match
                                subtractvalue = prevdata[k][xcol(col[3])]
                                if subtractvalue is not "" and subtractvalue is not None:  # just in case of var type error
                                    calc = calc - subtractvalue
                                    # print ("subtracting", subtractvalue)
                                # else:
                                #       print ("job",lookupjobnum,"col",j,": previous data did not have expected datatype")
                                # input()
                                trythisrowfirst = k
                                break
                    else:  # determine if we need to get from previous or current JSR. if no 3rd var, just subtract from ws1
                        subtractvalue = row[xcol(col[3])]  # subtract the second value
                        if subtractvalue is not "" and subtractvalue is not None:  # just in case of var type error
                            calc = calc - subtractvalue

                ws2.cell(row=i + 3, column=j).value = calc

    ###### here go calculations based on 995 using Key and MH Report.

    insertedkeyws = wb1.create_sheet(title="995 Key")  # create a blank worksheet in the JSR file
    insertedmhws = wb1.create_sheet(title="MH " + mhreporttimestamp)

    if CALC995 == "SKIPALL995":
        print("Skipping 995 calculations.")
        commenttext = "Skipped all 995 calculations, JSR ran in mode 1"
        ws2.cell(row=3, column=2).comment = Comment(commenttext, "JMai")
        pass
    elif CALC995 in ["SKIP995MONTH", "NORMAL"]:
        print("Copying 995 Key and MH Report")
        copyworksheet(wskey, insertedkeyws)  # copy data from the other worksheet
        copyworksheet(wsmh, insertedmhws, copyformatting=False)
        # add MH subtotal column in insertedmhws, this is needed for a static MH location so formulas don't break
        insertedkeyws.insert_cols(16)
        insertedkeyws.cell(row=1, column=16).value = "MH Subtotal"

        for cell in insertedmhws[1]:  # formats with colors and wraptext
            # start_color is background color, end_color is font color
            cell.fill = PatternFill(start_color='808080',fill_type="solid")
            cell.alignment = Alignment(horizontal='left', wrap_text=True)
            cell.font = Font(bold=True)

        # for each row in the main worksheet
        # get the job number
        # and search in the MH report col D for the YTD MH in Col J
        print("Starting 995 calculations")

        # prevent hard-coded column references. some are still hard coded like R and J
        ws2_costcenter_col = get_col_from_header_name(ws2, "Cost Cntr Home")
        ws2_jobnum_col = get_col_from_header_name(ws2, "Job #")
        ws2_YTDMH_col = get_col_from_header_name(ws2, "YTD Hourly Manhours", 3, exact=False)
        ws2_accrual_col = get_col_from_header_name(ws2, "Est Accruals for 995 and T&D")
        ws2_actmocost_col = get_col_from_header_name(ws2, "Actual Monthly Cost")
        ws2_acttotcost_col = get_col_from_header_name(ws2, "Actual Total Cost")
        ws2_mocost995_col = get_col_from_header_name(ws2, "Est Monthly Cost w 995 & T&D")
        ws2_actcostw995_col = get_col_from_header_name(ws2, "Est Total Cost w 995 & T&D")
        ws2_actmargw995_col = get_col_from_header_name(ws2, "Current % Margin w Est 995 & T&D")
        ws2_billings_col = get_col_from_header_name(ws2, "Total Billings")

        wsmh_busunit_col = get_col_from_header_name(insertedmhws, "COMPANY", 1)
        wsmh_YTDMH_col = get_col_from_header_name(insertedmhws, "CUMULATIVE", 1, exact=False)
        wsmh_allocation_col = get_col_from_header_name(insertedmhws, "ALLOCATION AREA", 1)

        wskey_costcenter_col = get_col_from_header_name(insertedkeyws, "Cost Cntr Home", 1)
        wskey_995rate_col = get_col_from_header_name(insertedkeyws, "995 Submitted", 1, exact=False)
        wskey_TNDrate_col = get_col_from_header_name(insertedkeyws, "T&D Rate", 1)
        wskey_MHsubtotal_col = get_col_from_header_name(insertedkeyws, "MH Subtotal", 1)

        list_of_995_numbers = [
            ["6008995", "250"],
            ["6009995", "277"],
            ["6010995", "251"],
            ["6012995", "252"],
            ["6013995", "253"],
            ["6014995", "254"],
            ["6015995", "255"],
            ["6017995", "256"],
            ["6018995", "257"],
            ["6019995", "258"],
            ["6020995", "259"],
            ["6021995", "260"],
            ["6022995", "261"],
            ["6023995", "262"],
            ["6024995", "263"],
            ["6025995", "264"],
            ["6026995", "265"],
            ["6027995", "266"],
            ["6028995", "267"],
            ["6029995", "268"],
            ["6031995", "269"],
            ["6032995", "270"],
            # ["6033995","271"], Water treatment is ignored
            ["6040995", "272"],
            ["6041995", "273"],
            ["6401995", "274"],
            ["6402995", "275"],
            ["6403995", "276"], ]

        if CALC995 == "SKIP995MONTH":
            commenttext = "Skipped partial 995 calculations, JSR ran in mode 2"
            ws2.cell(row=3, column=ws2_mocost995_col).comment = Comment(commenttext, "JMai")
            ws2.cell(row=3, column=ws2_actcostw995_col).comment = Comment(commenttext, "JMai")
            ws2.cell(row=3, column=ws2_actmargw995_col).comment = Comment(commenttext, "JMai")

        for j in range(4, ws2.max_row):
            if j % 500 == 0:
                print("... on row", j, "of", ws2.max_row)  # progress bar

            jobnumber = ws2.cell(row=j, column=ws2_jobnum_col).value  # load job number for this row
            if jobnumber:
                manhours = False  # Identifies if this has any MH in it, put # manhours into this
                areasupervision = False  # Identifies if this is 995 number, allocation area code into this

                ws2costcenter = ws2.cell(row=j, column=ws2_costcenter_col).value
                # this is the cost center for the current job number ie " REDLANDS - WA"

                for item in list_of_995_numbers:
                    if item[0] == jobnumber:
                        areasupervision = item[1]
                        break
                # calculate column R
                manhourint = 0
                for i in range(1, insertedmhws.max_row + 1):

                    # this if-else finds and enters the manhours
                    if areasupervision:
                        # find manhours as region supervision number.
                        mhreport_allocation = insertedmhws.cell(row=i, column=wsmh_allocation_col).value
                        if mhreport_allocation == "Total " + areasupervision:
                            manhourcell = insertedmhws.cell(row=i, column=wsmh_YTDMH_col)
                            manhourint = manhourint + manhourcell.value
                            newmanhourcell = ws2.cell(row=j, column=ws2_YTDMH_col)
                            newmanhourcell.value = -manhourint
                            manhours = True

                    else:
                        # find manhours as normal job number.
                        mhreport_jobnumber = insertedmhws.cell(row=i, column=wsmh_busunit_col).value
                        if mhreport_jobnumber == "Total Business Unit " + jobnumber:
                            manhourcell = insertedmhws.cell(row=i, column=wsmh_YTDMH_col)
                            manhours = manhourcell.value
                            newmanhourcell = ws2.cell(row=j, column=ws2_YTDMH_col)
                            newmanhourcell.value = manhours
                            break

                if areasupervision and manhours:
                    # write manhours into 995k sheet for static reference in formula
                    for x in range(1, insertedkeyws.max_row + 1):
                        key_costcenter = insertedkeyws.cell(row=x, column=wskey_costcenter_col).value
                        if key_costcenter == ws2costcenter:
                            #print("key_costcenter=", key_costcenter," -manhourint=", -manhourint)
                            #print("wskey_MHsubtotal_col",wskey_MHsubtotal_col)
                            #print("cellvalue",insertedkeyws.cell(row=x, column=wskey_MHsubtotal_col).value)
                            insertedkeyws.cell(row=x, column=wskey_MHsubtotal_col).value = -manhourint


                if "WATER TREATMENT" in ws2costcenter:
                    ws2.cell(row=j, column=ws2_accrual_col).value = "N/A"
                    ws2.cell(row=j, column=ws2_mocost995_col).value = "N/A"
                    ws2.cell(row=j, column=ws2_actcostw995_col).value = "N/A"
                    ws2.cell(row=j, column=ws2_actmargw995_col).value = "N/A"
                    ws2.cell(row=j, column=ws2_accrual_col).font = Font(color="808080")
                    ws2.cell(row=j, column=ws2_mocost995_col).font = Font(color="808080")
                    ws2.cell(row=j, column=ws2_actcostw995_col).font = Font(color="808080")
                    ws2.cell(row=j, column=ws2_actmargw995_col).font = Font(color="808080")
                    commenttext = "Skipped Calculation, Water Treatment not subject to 995"
                    ws2.cell(row=j, column=ws2_accrual_col).comment = Comment(commenttext, "JMai")
                else:

                    temp_rate995 = "NOT SET"    #This is the T&D + 995 rate to be used for this job number. this can vary

                    if CALC995 == "SKIP995MONTH":
                        ws2.cell(row=j, column=ws2_mocost995_col).value = "N/A"
                        ws2.cell(row=j, column=ws2_actcostw995_col).value = "N/A"
                        ws2.cell(row=j, column=ws2_actmargw995_col).value = "N/A"
                        ws2.cell(row=j, column=ws2_mocost995_col).font = Font(color="808080")
                        ws2.cell(row=j, column=ws2_actcostw995_col).font = Font(color="808080")
                        ws2.cell(row=j, column=ws2_actmargw995_col).font = Font(color="808080")

                    if manhours:
                        # There are manhours, so calculate the 995 costs in col s, by multiplying it against the area's key rate
                        for x in range(1, insertedkeyws.max_row + 1):
                            key_costcenter = insertedkeyws.cell(row=x, column=wskey_costcenter_col).value
                            if key_costcenter == ws2costcenter:
                                rate995 = insertedkeyws.cell(row=x, column=wskey_995rate_col)
                                rateTND = insertedkeyws.cell(row=x, column=wskey_TNDrate_col)
                                mhsubtotal =insertedkeyws.cell(row=x, column=wskey_MHsubtotal_col)

                                # Est accruals calculation
                                try:
                                    if areasupervision:
                                        temp_rate995 = rate995.value  # do not include TND for areasupervision # negative numbers for areasupervision codes
                                        accrual_formula = "='" + insertedkeyws.title + "'!" + mhsubtotal.coordinate + "*'" + insertedkeyws.title + "'!" + rate995.coordinate
                                        calculatedaccrual = temp_rate995 * mhsubtotal.value

                                    else:
                                        temp_rate995 = rate995.value + rateTND.value
                                        accrual_formula = "='" + insertedmhws.title + "'!" + manhourcell.coordinate + "*('" + insertedkeyws.title + "'!" + rate995.coordinate + "+ '"+ insertedkeyws.title + "'!" + rateTND.coordinate + ")"
                                        #accrual_formula = "='" + insertedmhws.title + "'!" + manhourcell.coordinate + "*('995 Key'!" + rate995.coordinate + "+ '995 Key'!" + rateTND.coordinate + ")"

                                        calculatedaccrual = temp_rate995 * manhourcell.value

                                except:
                                    accrual_formula = "ERROR"
                                # Set 995 cost
                                ws2.cell(row=j, column=ws2_accrual_col).value = accrual_formula

                                if CALC995 == "NORMAL":
                                    # perform est monthly cost w 995
                                    ws2.cell(row=j, column=ws2_mocost995_col).value = ws2.cell(row=j,column=ws2_actmocost_col).value + calculatedaccrual

                                    # calculate est total cost w 995
                                    actualtotcostw995 = calculatedaccrual + ws2.cell(row=j,column=ws2_acttotcost_col).value
                                    ws2.cell(row=j, column=ws2_actcostw995_col).value = actualtotcostw995

                                    if not areasupervision:
                                        # Set total marg w 995
                                        if ws2.cell(row=j, column=ws2_billings_col).value == 0:
                                            actualtotmarginw995 = 0
                                        else:
                                            actualtotmarginw995 = 1 - actualtotcostw995 / ws2.cell(row=j,column=ws2_billings_col).value
                                        ws2.cell(row=j, column=ws2_actmargw995_col).value = actualtotmarginw995

                                break

                            if x == insertedkeyws.max_row + 1:
                                print("failed to find any region code for job number ", jobnumber)

                    else:   # no manhours were found. no accrual, no manhour
                        ws2.cell(row=j, column=ws2_YTDMH_col).value = 0
                        ws2.cell(row=j, column=ws2_accrual_col).value = 0
                        ws2.cell(row=j, column=ws2_actmargw995_col).value = 0

    ####### end of 995 calculations

    ############################################################################
    ######### CONDITIONAL FORMATTING AND SPECIALTY CALCULATIONS GO HERE ########
    ############################################################################
    #
    # This is where we do special rules for worksheet modified_all
    #

    print("Adding conditional formatting and special if-then calculations")
    for i, row in enumerate(ws2.iter_rows(), 1):
        if i <= 3:  # skip the first 3 rows
            continue
        mark_if_actual_cost_is_greater_than_forecasted_cost(row)
        check_forecast_margin_change(row)
        clean_sales_vs_billings_values(row)
        mark_large_POC_receivables(row)
        mark_billings_over_contract_value(row)
        mark_actual_cost_over_billings_by_a_lot(row)
        add_number_formatting(row, HEADERMAP)

    #
    #
    #
    #
    ##########################################################################
    ## END OF CONDITIONAL FORMATTING AND SPECIALTY CALCULATIONS###############
    ##########################################################################

    print("Now splitting formatted spreadsheet into regional sheets")

    # split data from newly mapped sheet into multiple regional sheets
    # col_region = get_col_from_header_name(ws2, "Area")-1
    col_costcntr = get_col_from_header_name(ws2, "Cost Cntr Home") - 1

    for i, row in enumerate(ws2.iter_rows(), 1):
        if i <= 3:  # skip the first 3 rows
            continue
        # if row[col_region].value is None:  # Col AV contains the region code
        if row[col_costcntr].value is None:  # Col AX contains the region code
            print(row, "did not have a col region value")
            continue
        if i % 500 == 0:
            print("... on row", i, "of", ws2.max_row)

        putinother = True
        for worksheet in regional_worksheet_list:
            # if worksheet.code in row[col_region].value:  # copy to regional sheet if code matches
            if any(s in row[col_costcntr].value for s in worksheet.code):  # copy to regional sheet if code matches
                nextrow = worksheet.body.max_row + 1
                for cell in row:
                    if cell.col_idx > 65:
                        break  # don't do this past column 65, wasteful.
                    new_cell = worksheet.body.cell(row=nextrow, column=cell.col_idx, value=cell.value)
                    if cell.has_style:
                        new_cell.border = copy(cell.border)
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.comment = copy(cell.comment)
                        # new_cell.protection = copy(cell.protection)
                        # new_cell.alignment = copy(cell.alignment)
                putinother = False
                break
        if putinother is True:  # copy to Other page if no match
            nextrow = wsother.max_row + 1
            # print (row[col_region].value,"- Other")
            for cell in row:
                if cell.col_idx > 65:
                    break  # don't do this past column 65, wasteful.
                new_cell = wsother.cell(row=nextrow, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell.border = copy(cell.border)
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.comment = copy(cell.comment)

    print("Now preparing highlighted rows & styling")

    # this takes a while to execute so hide comment it out when running if you're not specifically testing it
    countdown = len(all_modified_worksheet_list)
    for sheet in all_modified_worksheet_list:
        sys.stdout.write(str(countdown))
        sys.stdout.flush()
        # print(countdown, end="")
        # print(countdown,sheet.title,end="")
        countdown = countdown - 1
        # print("Highlighting alternate rows")
        if not DEBUG:
            highlight_alternate_rows(sheet)  # option to skip this during debug mode

        # print("Hiding TBD columns")
        for col in ['B', 'AW', 'AY', 'BE']:
            sheet.column_dimensions[col].hidden = True

        # hide RSTUV column if 995 was skipped
        if CALC995 == "SKIPALL995":
            for col in ['R', 'S', 'T', 'U', 'V']:
                sheet.column_dimensions[col].hidden = True

        # hide row 2. Row 1 is left for old header titles .
        for row in [1]:
            sheet.row_dimensions[row].hidden = True

        # freezing pane
        sheet.freeze_panes = sheet['A4']

        # print("Adjusting Column Widths")
        for i in range(48, 59):
            sheet.column_dimensions[get_column_letter(i)].width = 16
        for i in list(range(6, 47)) + list(range(53, 56)):
            sheet.column_dimensions[get_column_letter(i)].width = 12
        sheet.column_dimensions['E'].width = 35

        # print("Putting Filter in Place")
        FullRange = "A3:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
        sheet.auto_filter.ref = FullRange

        # conditional formatting to highlight subtotals
        # this finds rows with "Total" in column A and has an equation in col F, which is pretty unique to subtotal rows
        # subtotals will be activated during macro portion because openpyxl does not support it.
        # this conditional formatting doesn't work until needs user (or macro) goes in and click "apply formatting".
        sheet.conditional_formatting.add("A1:BG5000",
                                         FormulaRule(formula=['=AND(ISNUMBER(SEARCH(" Total",$A1)),ISFORMULA($F1))'],
                                                     stopIfTrue=True,
                                                     font=Font(color="000000"),
                                                     fill=PatternFill(bgColor="aaaaaa"))
                                         )
        # cant use conditional formatting or else it destroys red-flagged cells

    # move original sheet to end
    move_sheet(wb1, 0, len(wb1._sheets) - 1)

    # format and copy over prev month sheet
    prevworksheet = wb1.create_sheet(title="Original_Prev_All")
    for row in wsprev.iter_rows():
        prevworksheet.append(cell.value for cell in row)
    for cell in prevworksheet[3]:  # formats with colors and wraptext
        cell.fill = PatternFill(start_color='5B9BD5',
                                fill_type="solid")  # start_color is background color, end_color is font color
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(color="FFFFFF", bold=True)

    prevworksheet['B2'].value = oldJSRtimestamp

    ws2.active = 0
    ws1.views.sheetView[0].tabSelected = False

    # create save file name based on todays year-month-date
    # this is the filename you save the formatted JSR as
    YMD = datetime.datetime.now().strftime("%y%m%d")
    HMS = datetime.datetime.now().strftime("%H%M%S")

    save_directory = output_dir + "/Reports " + datetime.datetime.now().strftime("%Y-%m-%d") + "/"  # ends in /"
    save_file_name = "JSR " + newJSRtimestamp + " vs " + oldJSRtimestamp + "(" + YMD + "-" + HMS + ")" + ".xlsx"
    savefile = save_directory + save_file_name

    print("")
    print("Excel processing finished.")

    try:
        if not os.path.isdir(save_directory):
            os.makedirs(save_directory)
        print("Attempting to save to: ")
        print(savefile)
        wb1.save(savefile)
        print("Save complete")
    except KeyboardInterrupt:
        raise KeyboardInterrupt
    except:
        print("----------------------------------------------------------")
        print("ERROR: Could not save. Close the destination file and then")
        input("Press 'ENTER' to try saving again")
        try:
            print("Trying last time to save... ", end="")
            wb1.save(savefile)
            print("Save complete")
        except KeyboardInterrupt:
            raise KeyboardInterrupt
        except:
            print("ERROR: Still could not save. Terminating script.")
            sys.exit(0)

    print()
    AUTOMATE_EXCEL_FORMATTING(savefile, save_file_name)
    input("Script complete. Press 'ENTER' to close")


def trytoloadworkbook(address):
    print("Loading:", address, end="     ")
    try:
        a = load_workbook(address)
    except KeyboardInterrupt:
        raise KeyboardInterrupt
    except:
        print("Loading Error. Closing this file in excel will likely fix the problem.")
        input("Press 'ENTER' to close")
        sys.exit()


    print(" Done.")
    return a


def copyworksheet(source, destination, copyformatting=True):
    for row in source:
        for cell in row:
            destination[cell.coordinate].value = cell.value
            if copyformatting:
                destination[cell.coordinate].fill = copy(cell.fill)
                destination[cell.coordinate].number_format = copy(cell.number_format)
                destination[cell.coordinate].alignment = copy(cell.alignment)
                destination[cell.coordinate].font = copy(cell.font)
                # destination[cell.coordinate].comment = copy(cell.comment)

    for idx, rd in source.row_dimensions.items():
        destination.row_dimensions[idx] = copy(rd)
    for idx, rd in source.column_dimensions.items():
        destination.column_dimensions[idx] = copy(rd)


def newsheetwithheaders(workbook, sheettitle, headermap, mhtimestamp, color=""):
    worksheet = workbook.create_sheet(title=sheettitle)  # create new sheet with region title

    newheaders = []
    oldheaders = []  # names of headers from old JSR, for reference in top row
    oldws = workbook["Original_All"]
    for col in headermap:  # extracts and creates list of header names from headermap
        if col[0] == "YTD Hourly Manhours":
            newheaders.append(
                col[0] + " " + mhtimestamp)  # special header naming for manhours to show which week end it is
        else:
            newheaders.append(col[0])
        if col[2] != "" and col[3] == "":
            oldheaders.append(oldws[col[2] + '3'].value)
        else:
            oldheaders.append("")

    worksheet.append(list())  # add 2 blank rows
    worksheet.append(oldheaders)
    worksheet.append(newheaders)  # adds header names to worksheet

    for cell in worksheet[3]:  # formats with colors and wraptext
        cell.fill = PatternFill(start_color='5B9BD5',
                                fill_type="solid")  # start_color is background color, end_color is font color
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(color="FFFFFF", bold=True)
        if get_column_letter(cell.col_idx) in ['R', 'S', 'T', 'U', 'V']:
            cell.fill = PatternFill(start_color='808080',
                                    fill_type="solid")  # start_color is background color, end_color is font color
    # color tabs
    if color != "" and type(color) == str:
        if color.lower() == 'green':
            worksheet.sheet_properties.tabColor = 'bdf1d8'
        elif color.lower() == 'blue':
            worksheet.sheet_properties.tabColor = 'bdd7ee'
        elif color.lower() == 'red':
            worksheet.sheet_properties.tabColor = 'fac2c2'
        elif color.lower() == 'yellow':
            worksheet.sheet_properties.tabColor = 'f5ff99'
        elif color.lower() == 'purple':
            worksheet.sheet_properties.tabColor = 'd3bcf2'
        elif color.lower() == 'orange':
            worksheet.sheet_properties.tabColor = 'f3d8bb'
        else:
            worksheet.sheet_properties.tabColor = color

    return worksheet


class newWorksheet:
    def __init__(self, title, code, workbook, mhtimestamp, color=""):
        self.title = title
        self.code = code
        self.body = newsheetwithheaders(workbook, title, HEADERMAP, mhtimestamp, color)


def move_sheet(wb, from_loc=None, to_loc=None):
    sheets = wb._sheets

    # if no from_loc given, assume last sheet
    if from_loc is None:
        from_loc = len(sheets) - 1

    # if no to_loc given, assume first
    if to_loc is None:
        to_loc = 0

    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)


def highlight_alternate_rows(worksheet):
    blue_borderstyle = Border(top=Side(style='thin', color="9BC2E6"), bottom=Side(style='thin', color="9BC2E6"))
    blue_fillstyle = PatternFill(start_color='ddebf7', fill_type="solid")
    gray_borderstyle = Border(top=Side(style='thin', color="bfbfbf"), bottom=Side(style='thin', color="bfbfbf"))
    gray_fillstyle = PatternFill(start_color='d9d9d9', fill_type="solid")

    # for i in range (5,worksheet.max_row,2):    # skip first 3(or 4) lines
    # worksheet.row_dimensions[i].fill = blue_fillstyle
    # worksheet.row_dimensions[i].border = blue_borderstyle

    for i in range(4, worksheet.max_row + 1):  # go through every row, every cell. start on line 4

        if i % 100 == 0:    # progress bar
            sys.stdout.write(".")
            sys.stdout.flush()
            #print(".", end="")

        for cell in worksheet[i]:

            if cell.col_idx > 60:
                break  # don't go past column 60, wasteful.

            if get_column_letter(cell.col_idx) in ['R', 'S', 'T', 'U', 'V']:  # special highlighting for 995 columns
                cell.fill = gray_fillstyle
                cell.border = gray_borderstyle
            elif i % 2 == 1:  # highlight cells on odd rows
                if cell.fill.start_color.rgb is "00000000":  # check if cell color is default, this step is necessary to avoid filling red-flagged cells
                    cell.fill = blue_fillstyle
                    cell.border = blue_borderstyle


            #ws2_billings_col = get_col_from_header_name(ws2, "Total Billings")
            #if cell.value == 0 and cell.column != ws2_billings_col:  # apply gray formatting to zeros, except in billing column Y and MH col T
            if cell.value == 0 and cell.column not in ['L']:  # apply gray formatting to zeros, except in billing column L
                cell.font = Font(color="b2b2b2")


def killer():
    print("killer!")
    signal.CTRL_C_EVENT
    raise KeyboardInterrupt


def xcol(alphanumeric):
    return column_index_from_string(alphanumeric) - 1


def newest_file(path, keyword=""):
    # This function returns the most recently updated file in directory "path"
    # Optional, add a filter keyword filenames

    files = os.listdir(path)
    if files == []:
        print("Error: There are no files in", path)
        print("Please put JSR files into the input folders and try again.")
        input("Press 'ENTER' to quit")
        sys.exit()

    # search for a keyword.

    tracksheets = []
    for basename in files:
        if basename.lower().find(keyword.lower()) != -1:
            # print (basename)
            # print (basename.replace('\\','/'))
            tracksheets.append(basename)

    paths = [os.path.join(path, basename) for basename in tracksheets]
    # print ("paths = ",paths)
    if paths:  # if paths is not empty
        return max(paths, key=os.path.getctime).replace('\\', '/')
    else:
        print("Failed to load file at ", path)
        return False

def parse_for_date(searchterm, key):
    return searchterm[searchterm.find(key) + len(key):]


def get_timestamp_str(filepath, worksheet):
    # Compares 3 timestamps and returns the earliest one
    #    1. date in cell BP4
    #    2. date file last modified
    #    3. today's date
    # Returns string, ex: '2019-12-31'

    strdatefromfilemod = time.strftime('%Y-%m-%d', time.gmtime(os.path.getmtime(filepath)))
    strdatefromexcel = parse_for_date(searchterm=worksheet['BP4'].value, key="||")
    strdatestampnow = datetime.datetime.now().strftime("%Y-%m-%d")

    a = min(strdatefromfilemod, strdatefromexcel, strdatestampnow)
    # print(a, "appears to be the timestamp for", filepath)
    return a


Redfillstyle = PatternFill(start_color='FF7D7D', fill_type="solid")
Pinkfillstyle = PatternFill(start_color='FFAFAF', fill_type="solid")


def mark_if_actual_cost_is_greater_than_forecasted_cost(row):
    # Q is actual cost, N is Forecasted cost
    ACTCOSTCOL="Q"
    FORCOSTCOL="N"
    if row[xcol(ACTCOSTCOL)].value is not None and row[xcol(FORCOSTCOL)].value is not None:
        ActualGreaterThanForecast = row[xcol(ACTCOSTCOL)].value - row[xcol(FORCOSTCOL)].value
        if ActualGreaterThanForecast > 0:
            row[xcol(ACTCOSTCOL)].fill = Redfillstyle
            commenttext = "Actual Cost greater than Forecast Cost"
            row[xcol(ACTCOSTCOL)].comment = Comment(commenttext, "JMai")
        return True
    else:
        return False


def check_forecast_margin_change(row):
    # AP is Forecast margin monthly change
    FCSTCHANGECOL="AP"
    CONTRTYPECOL = "A"
    if row[xcol(FCSTCHANGECOL)].value is not None:
        if abs(row[xcol(FCSTCHANGECOL)].value) > 20000:
            if row[xcol(CONTRTYPECOL)].value in ["LJ", "FJ"]:
                row[xcol(FCSTCHANGECOL)].fill = Redfillstyle
                commenttext = "$20k+ margin change. Confirm costs are in correct cost buckets, or adjust E-1 forecast."
                row[xcol(FCSTCHANGECOL)].comment = Comment(commenttext, "JMai")
        return True
    else:
        return False

def clean_sales_vs_billings_values(row):
    # AE and AF is billings vs sales, delete if negative
    BVSCOL=["AE", "AF"]
    for a in BVSCOL:
        b = row[xcol(a)].value
        if b is not None:  # make sure we don't compare "" with an integer.
            if b < 0:
                row[xcol(a)].value = ""
    return


def mark_large_POC_receivables(row):
    # AB is POC receivables
    POCRECCOL="AB"
    POC_threshold = 20000  # ignore unless POC is over $20k
    a = row[xcol(POCRECCOL)].value
    if a:
        if a > POC_threshold:
            commenttext = "POC Receivables above $20k"
            row[xcol(POCRECCOL)].comment = Comment(commenttext, "JMai")
            row[xcol(POCRECCOL)].fill = Redfillstyle
            return True
        elif a < -POC_threshold:
            commenttext = "POC Receivables below -$20k"
            row[xcol(POCRECCOL)].comment = Comment(commenttext, "JMai")
            row[xcol(POCRECCOL)].fill = Redfillstyle
            return True
    return False


def mark_billings_over_contract_value(row):
    # Two levels of highlight. Light Red level 1, Dark Red level 2
    # Level 1: mark if est sales < actual cost * 1.22
    # Level 2: mark if est sales < billings (This one is more important, highlight this one if you have to choose)
    threshold = 100

    # L is billings, F is contract Value, Q is actual Total Cost, A in contract Type
    BILLCOL="L"
    CONVALCOL="F"
    TOTCOSTCOL="Q"
    CONTRTYPECOL="A"
    if row[xcol(BILLCOL)].value is not None and row[xcol(CONVALCOL)].value is not None:
        if row[xcol(CONVALCOL)].value > 5:
            # ignore if contract value is tiny
            if row[xcol(BILLCOL)].value > row[xcol(CONVALCOL)].value + threshold:
                commenttext = "Billing exceed Contract Value. Change order needed if there are more billings."
                row[xcol(CONVALCOL)].comment = Comment(commenttext, "JMai")
                row[xcol(CONVALCOL)].fill = Redfillstyle
                return True
            elif row[xcol(TOTCOSTCOL)].value * 1.22 > row[xcol(CONVALCOL)].value and row[xcol(TOTCOSTCOL)].value > row[
                xcol(CONVALCOL)].value + threshold:
                # apply level 2 only if it is a CJ job
                if row[xcol(CONTRTYPECOL)].value == "CJ":
                    commenttext = "Revenue accrual (based on actual costs) is below 18% margin. Possible CO needed."
                    row[xcol(CONVALCOL)].comment = Comment(commenttext, "JMai")
                    row[xcol(CONVALCOL)].fill = Pinkfillstyle
                    return True
    else:
        return False


def mark_actual_cost_over_billings_by_a_lot(row):
    # Two levels of highlight. Light Red level 1, Dark Red level 2
    # Level 1: mark if cost * 125% > billings
    # Level 2: mark if cost > billings + 15000 (This one is more important, highlight this one if you have to choose)

    # L is billings,       Q is actual Total Cost
    BILLCOSTCOL="L"
    TOTCOSTCOL="Q"
    Cost_Threshold = 15000  # ignore unless cost is over billings by a lot, otherwise everything gets flagged
    Cost_Perc_Threshold = 1.25

    if row[xcol(BILLCOSTCOL)].value is not None and row[xcol(TOTCOSTCOL)].value is not None:
        # if zerotrigger: print("debug 1,", end="")
        if row[xcol(TOTCOSTCOL)].value > 3000:  # ignore if cost is tiny, less than 3000
            #       if zerotrigger: print("2", end="")
            if row[xcol(TOTCOSTCOL)].value > row[xcol(BILLCOSTCOL)].value + Cost_Threshold:
                #              if zerotrigger: print("3", end="")
                commenttext = "Actual Cost exceeds Total Billings by over $15k"
                row[xcol(BILLCOSTCOL)].comment = Comment(commenttext, "JMai")
                row[xcol(BILLCOSTCOL)].fill = Redfillstyle
                return True
            elif row[xcol(TOTCOSTCOL)].value * Cost_Perc_Threshold > row[xcol(BILLCOSTCOL)].value:
                #             if zerotrigger: print("4", end="")
                commenttext = "Billings to Cost Ratio below 1.25"
                row[xcol(BILLCOSTCOL)].comment = Comment(commenttext, "JMai")
                row[xcol(BILLCOSTCOL)].fill = Pinkfillstyle
                return True
        #    if zerotrigger: print("5", end="")
    return False


def add_number_formatting(row, headermap):
    for i, cell in enumerate(row):
        cell.number_format = headermap[i][1]


def get_col_from_header_name(ws, headername, header_row=3, exact=True):
    # checks a ws in a particular row for a particular string. returns alpha column
    # exact true means a hard match.
    # exact false means it can contain the search string
    for cell in ws[header_row]:

        try:
            value = cell.value.lower()
        except:
            value = cell.value

        if exact:
            if headername.lower() == value:
                return cell.col_idx
        else:
            if headername.lower() in value:
                return cell.col_idx


###############################################
# This executes main script
if __name__ == '__main__':
    # try: main()
    main()
#        except (KeyboardInterrupt):
# print ('Script interrupted by keyboard press')
# input ("Press enter to close script")
# sys.exit()
